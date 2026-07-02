

from math import radians, tan
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from parapy.core import action
import openpyxl
from Optimizer.Optim_modules.propeller_weight import run_model as compute_prop_weight

from parapy.geom import (GeomBase, translate, Point, FittedCurve, rotate,
                         XOY, TransformedCurve, ScaledCurve,
                         LoftedSurface, LoftedSolid, RotatedCurve, Cylinder)
from parapy.core import Base, Part, Input, Attribute, child
from parapy.geom import Position
from parapy.geom import Cylinder, Box, Position, Fused
import math

def _read_blade_table(filepath: str):

    df = pd.read_excel(filepath, header=None)
    df = df.dropna(how='all').reset_index(drop=True)

    if df.shape[1] < 3:
        raise ValueError(f"Excel file must have at least 3 columns; got {df.shape[1]}.")
    if not pd.api.types.is_numeric_dtype(df.iloc[0, 0]):
        df = df.iloc[1:].reset_index(drop=True)

    df = df.iloc[:, :3].astype(float)
    df.columns = ['r_R', 'c_ctip', 'twist']

    if df['r_R'].min() < 0 or df['r_R'].max() > 1:
        raise ValueError("r/R_tip values must be in [0, 1].")
    if (df['c_ctip'] <= 0).any():
        raise ValueError("c/c_tip values must all be positive.")

    df = df.sort_values('r_R').reset_index(drop=True)
    return list(df.itertuples(index=False, name=None))



class BladeSection(GeomBase):
    airfoil_curve: object = Input()
    chord: float = Input()
    twist_deg: float = Input()
    span_y: float = Input()
    sweep_x: float = Input()
    mesh_deflection: float = Input(1e-4)

    @Part
    def _unit_positioned(self):
        return TransformedCurve(
            curve_in=self.airfoil_curve,
            from_position=rotate(translate(XOY, 'x', 1), 'x', -90, deg=True),
            to_position=translate(self.position,
                                  'y', self.span_y,
                                  'x', self.sweep_x),
            hidden=True
        )

    @Part
    def _scaled(self):
        return ScaledCurve(
            curve_in=self._unit_positioned,
            reference_point=self._unit_positioned.start,
            factor=self.chord,
            mesh_deflection=self.mesh_deflection,
            hidden=True
        )

    @Part
    def curve(self):
        return RotatedCurve(
            curve_in=self._scaled,
            rotation_point=self._scaled.start,
            vector=self.position.Vy,
            angle=radians(self.twist_deg),
            mesh_deflection=self.mesh_deflection
        )



class Blade(GeomBase):


    R_tip: float = Input()
    c_tip: float = Input()
    hub_radius: float = Input(0.0)
    sweep_TE: float = Input(0.0)
    airfoil_file: str = Input()
    blade_data_file_horizontal: str = Input()
    mesh_deflection: float = Input(1e-4)

    @Attribute
    def blade_table(self):
        return _read_blade_table(self.blade_data_file_horizontal)

    @Attribute
    def n_sections(self):
        return len(self.blade_table)

    @Attribute
    def _airfoil_pts(self):
        with open(self.airfoil_file, 'r') as f:
            pts = []
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    x, y = line.split(None, 1)
                    pts.append(Point(float(x), float(y)))
                except ValueError:
                    continue
        return pts

    @Part
    def _base_airfoil(self):
        return FittedCurve(points=self._airfoil_pts,
                           mesh_deflection=self.mesh_deflection,
                           hidden=True)

    @Part
    def sections(self):
        return BladeSection(
            quantify=self.n_sections,
            airfoil_curve=self._base_airfoil,
            chord=self.c_tip * self.blade_table[child.index][1],
            twist_deg=self.blade_table[child.index][2],
            span_y=self.hub_radius + self.R_tip * self.blade_table[child.index][0],  # <-- offset added
            sweep_x=((self.hub_radius + self.R_tip * self.blade_table[child.index][0])
                     * tan(radians(self.sweep_TE))),  # <-- sweep also offset consistently
            mesh_deflection=self.mesh_deflection,
            position=self.position
        )

    @Attribute
    def _section_curves(self):
        return [sec.curve for sec in self.sections]

    @Part
    def blade_surface(self):
        return LoftedSurface(
            profiles=self._section_curves,
            mesh_deflection=self.mesh_deflection
        )

    @Part
    def blade_solid(self):
        return LoftedSolid(
            profiles=self._section_curves,
            mesh_deflection=self.mesh_deflection
        )



class Propeller(GeomBase):

    hub_radius: float = Input(0.08)
    R_tip: float = Input(1.0)
    c_tip:float = Input(0.2)
    horizontal_num_blades: int = Input(3)
    sweep_TE: float = Input(0.0)
    airfoil_file: str = Input()
    blade_data_file_horizontal: str = Input()
    mesh_deflection: float = Input(1e-4)

    @Part
    def hub(self):

        return Cylinder(
            radius=self.hub_radius,
            height=self.c_tip,
            centered=True,
            color='red',
            position=self.position.rotate90('x' if child.index in (2, 3) else 'y')
        )

    @Part
    def blades(self):
        return Blade(
            quantify=self.horizontal_num_blades,
            R_tip=self.R_tip,
            c_tip=self.c_tip,
            hub_radius=self.hub_radius,
            sweep_TE=self.sweep_TE,
            airfoil_file=self.airfoil_file,
            blade_data_file_horizontal=self.blade_data_file_horizontal,
            mesh_deflection=self.mesh_deflection,
            position=rotate(
                self.position.translate("x", self.c_tip/2),
                'x' if child.index in (2, 3) else 'x',
                child.index * 360.0 / self.horizontal_num_blades,
                deg=True
            ),
            color='orange',
            label=f'blade_{child.index}'
        )

    @Attribute
    def shape(self):
        return Fused(
            shape_in=self.hub,
            tool=[b.blade_solid for b in self.blades]
        )


class HorizontalPropulsion(Base):

    horizontal_propulsor_loc = Input(0)
    radius: float = Input(1.0)
    weight_radius: float = Input(None)
    envelope_radius: float = Input()
    position: object = Input()
    strut_size: float = Input()
    horizontal_num_blades: int = Input(3)
    #horizontal_prop_weight: float = Input(20)
    hub_to_tip_h = Input(0.2)

    c_tip: float = Input()
    sweep_TE: float = Input(0.0)
    airfoil_file: str = Input()
    blade_data_file_horizontal: str = Input()
    mesh_deflection: float = Input(1e-4)
    design_airspeed: float = Input(9)
    blade_data_version: int = Input(0)

    @Attribute
    def radius_true(self):
        cross_section_radius = self.envelope_radius * math.sqrt(1 - self.horizontal_propulsor_loc ** 2)
        return cross_section_radius + self.strut_size

    @Attribute
    def weight(self):
        r = self.weight_radius if self.weight_radius is not None else self.radius
        result = compute_prop_weight({
            "excel_path": self.blade_data_file_horizontal,
            "prop_radius_m": r,
            "blade_count": self.horizontal_num_blades,
            "hub_to_tip": self.hub_to_tip_h,
        })
        return result["total_assembly_weight_kg"]

    @Part
    def propulsors(self):
        return Propeller(
            quantify=2,
            R_tip=self.radius,
            c_tip=self.c_tip,
            horizontal_num_blades=self.horizontal_num_blades,
            sweep_TE=self.sweep_TE,
            airfoil_file=self.airfoil_file,
            blade_data_file_horizontal=self.blade_data_file_horizontal,
            mesh_deflection=self.mesh_deflection,
            position=(
                self.position
                .translate(
                    "x",
                    self.radius_true if child.index == 0 else
                    -self.radius_true if child.index == 1 else 0
                )
                .rotate90('z')
                .translate(
                    "z",
                    -self.horizontal_propulsor_loc*self.envelope_radius if child.index in (0,1) else 0
                )
            )
        )

    PROP_EXCEL_PATH = r"C:\Users\Arnav\Desktop\Career\MSc Aerospace\KBE\parapy_tutorial\ProjectUpdated\KBE_Arnav_Arnesh_Assignment\SubmissionDocuments\KBE_Final\KBE\prop_Ct_Cp_vs_J.xlsx"

    def _write_to_prop_excel(self, j_vals, ct_vals=None, cp_vals=None):
        wb = openpyxl.load_workbook(self.PROP_EXCEL_PATH)
        ws = wb['Ct_Cp_vs_J']

        vals = ct_vals if ct_vals is not None else cp_vals
        col = 2 if ct_vals is not None else 3

        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=1).value = None
            ws.cell(row=row_idx, column=col).value = None

        for i, (j, v) in enumerate(zip(j_vals, vals), start=2):
            ws.cell(row=i, column=1).value = j
            ws.cell(row=i, column=col).value = v

        wb.save(self.PROP_EXCEL_PATH)


    @Attribute
    def shape(self):
        shapes = [p.shape for p in self.propulsors]
        return Fused(shape_in=shapes[0], tool=shapes[1:])

    @Attribute
    def _performance_data(self):
        import jpype
        import jpype.imports
        import numpy as np
        import env_Arnesh as env
        import os
        from math import pi

        _ = self.blade_data_version

        D = 2.0 * self.radius
        R = self.radius
        V = self.design_airspeed
        rho = 1.225
        kv = 1.5e-5#kinematic viscosity
        sos = 340.0 #speed of sound
        #these values should be calculated at cruise altitude and not hardcoded for next version
        spinner_dia = self.hub_to_tip_h * D


        if not jpype.isJVMStarted():
            jpype.startJVM(
                jpype.getDefaultJVMPath(),
                classpath=[
                    os.path.join(env.basepath, 'JavaProp.jar'),
                    os.path.join(env.basepath, 'MHClasses.jar'),
                ],
            )

        from MH.JavaProp import Propeller
        from MH.AeroTools.Airfoils import Airfoil

        def make_airfoil(n):
            af = Airfoil()
            af.Init(n)
            return af


        blade_sections = 40
        prop = Propeller(blade_sections)
        prop.Name = 'JavaProp-KBE-FromExcel'

        prop.Density = rho
        prop.KinematicViscosity = kv
        prop.SpeedOfSound = sos

        prop.removeAirfoils()
        prop.addAirfoil(jpype.JDouble(0.000), make_airfoil(13)) #hardcoded airfoil sections from example, need to be updated with xfoil analysis of .dat file provided by user
        prop.addAirfoil(jpype.JDouble(0.333), make_airfoil(13))
        prop.addAirfoil(jpype.JDouble(2 / 3), make_airfoil(12))
        prop.addAirfoil(jpype.JDouble(1.000), make_airfoil(10))

        prop.addAlfa(jpype.JDouble(0.00), jpype.JDouble(3.0))
        prop.addAlfa(jpype.JDouble(0.25), jpype.JDouble(3.0))
        prop.addAlfa(jpype.JDouble(0.50), jpype.JDouble(3.0))
        prop.addAlfa(jpype.JDouble(0.75), jpype.JDouble(3.0))
        prop.addAlfa(jpype.JDouble(1.00), jpype.JDouble(3.0))
        #Sets the design angle of attack to 3 degrees at five spanwise stations (0%, 25%, 50%, 75%, 100%).
        #All set to 3° here, meaning a uniform angle of attack along the whole blade. Needs to be read from
        #blade geometry file.

        prop.BladeCount = self.horizontal_num_blades
        prop.rRSpinner = jpype.JDouble(spinner_dia / D)
        prop.removeShroud()
        prop.hasSquareTips = 0

        prop.incrementBladeAngle(jpype.JDouble(0.0))
        prop.multiplyBladeAngle(jpype.JDouble(1.0))
        prop.incrementChord(jpype.JDouble(0.0))
        prop.multiplyChord(jpype.JDouble(1.0))
        prop.taperChord(jpype.JDouble(1.0))


        rpm_design = 3000.0
        Omega_design = 2.0 * pi * (rpm_design / 60.0)
        thrust_design = 100.0
        prop.performPropellerDesign(
            jpype.JDouble(V),
            jpype.JDouble(Omega_design),
            jpype.JDouble(R),
            jpype.JDouble(0.0),
            jpype.JDouble(thrust_design)
        )#Initialized


        table = _read_blade_table(self.blade_data_file_horizontal)

        r_R_excel = np.array([row[0] for row in table])
        c_R_excel = np.array([(row[1] * self.c_tip) / R for row in table])
        twist_excel = np.array([row[2] for row in table])

        n = len(table)


        JDoubleArray = jpype.JArray(jpype.JDouble)

        rR_java = JDoubleArray(n)
        cR_java = JDoubleArray(n)
        beta_java = JDoubleArray(n)

        for i in range(n):
            rR_java[i] = jpype.JDouble(r_R_excel[i])
            cR_java[i] = jpype.JDouble(c_R_excel[i])
            beta_java[i] = jpype.JDouble(twist_excel[i])


        prop.interpolateGeometry(rR_java, cR_java, beta_java, jpype.JInt(n))
        #replace initialized geometry with true geometry

        J_arr = list(np.linspace(0.05, 1.2, 60))
        Ct_arr = []
        Cp_arr = []

        for J in J_arr:
            n = V / (J * D)
            Omega = 2.0 * pi * n
            prop.performAnalysis(
                jpype.JDouble(V), jpype.JDouble(Omega), jpype.JDouble(R),
                jpype.JDouble(rho), jpype.JDouble(kv), jpype.JDouble(sos)
            )
            Ct_arr.append(max(0.0, float(prop.CT)))
            Cp_arr.append(max(0.0, float(prop.CP)))


        rpm_arr = [0.0] + list(np.arange(100, 7100, 100, dtype=float))
        T_arr = [0.0]
        eta_arr = [0.0]

        for rpm in rpm_arr[1:]:
            n = rpm / 60.0
            Omega = 2.0 * pi * n
            prop.performAnalysis(
                jpype.JDouble(V), jpype.JDouble(Omega), jpype.JDouble(R),
                jpype.JDouble(rho), jpype.JDouble(kv), jpype.JDouble(sos)
            )
            T_arr.append(max(0.0, float(prop.getThrust())))
            eta_arr.append(max(0.0, min(1.0, float(prop.Eta))))

        return {
            "J": J_arr,
            "Ct": Ct_arr,
            "Cp": Cp_arr,
            "rpm": rpm_arr,
            "thrust": T_arr,
            "eta": [e * 100.0 for e in eta_arr],
        }

    @action(label="Plot Ct vs J")
    def plot_ct_vs_j(self):
        data = self._performance_data
        fig, ax = plt.subplots(figsize=(7, 4))
        ax.plot(data["J"], data["Ct"], color="#1f77b4", linewidth=1.8)
        ax.set_xlabel("Advance ratio  J = V / (nD)", fontsize=11)
        ax.set_ylabel("$C_T$", fontsize=11)
        ax.set_title(
            f"Thrust coefficient  —  R = {self.radius} m, "
            f"{self.horizontal_num_blades} blades"
        )
        ax.grid(which="both", color="0.8", linestyle="-")
        ax.set_xlim(left=0)
        ax.set_ylim(bottom=0)
        fig.tight_layout()
        fig.show()
        self._write_to_prop_excel(data["J"], ct_vals=data["Ct"])
        return "Ct vs J plotted and written to Excel"

    @action(label="Plot Cp vs J")
    def plot_cp_vs_j(self):
        data = self._performance_data
        fig, ax = plt.subplots(figsize=(7, 4))
        ax.plot(data["J"], data["Cp"], color="#ff7f0e", linewidth=1.8)
        ax.set_xlabel("Advance ratio  J = V / (nD)", fontsize=11)
        ax.set_ylabel("$C_P$", fontsize=11)
        ax.set_title(
            f"Power coefficient  —  R = {self.radius} m, "
            f"{self.horizontal_num_blades} blades"
        )
        ax.grid(which="both", color="0.8", linestyle="-")
        ax.set_xlim(left=0)
        ax.set_ylim(bottom=0)
        fig.tight_layout()
        fig.show()
        self._write_to_prop_excel(data["J"], cp_vals=data["Cp"])
        return "Cp vs J plotted and written to Excel"

    @action(label="Plot Thrust vs RPM")
    def plot_thrust_vs_rpm(self):
        data = self._performance_data
        fig, ax = plt.subplots(figsize=(7, 4))
        ax.plot(data["rpm"], data["thrust"], color="#2ca02c", linewidth=1.8)
        ax.set_xlabel("RPM", fontsize=11)
        ax.set_ylabel("Thrust  (N)", fontsize=11)
        ax.set_title(
            f"Thrust vs RPM  at V = {self.design_airspeed} m/s  —  "
            f"R = {self.radius} m, {self.horizontal_num_blades} blades"
        )
        ax.set_xlim(0, 7000)
        ax.set_ylim(bottom=0)
        ax.grid(which="both", color="0.8", linestyle="-")
        fig.tight_layout()
        fig.show()
        return "Thrust vs RPM plotted"

    @action(label="Plot Efficiency vs RPM")
    def plot_efficiency_vs_rpm(self):
        data = self._performance_data
        fig, ax = plt.subplots(figsize=(7, 4))
        ax.plot(data["rpm"], data["eta"], color="#9467bd", linewidth=1.8)
        ax.set_xlabel("RPM", fontsize=11)
        ax.set_ylabel("Propulsive efficiency  η  (%)", fontsize=11)
        ax.set_title(
            f"Efficiency vs RPM  at V = {self.design_airspeed} m/s  —  "
            f"R = {self.radius} m, {self.horizontal_num_blades} blades"
        )
        ax.set_xlim(0, 7000)
        ax.set_ylim(0, 100)
        ax.grid(which="both", color="0.8", linestyle="-")
        fig.tight_layout()
        fig.show()
        return "Efficiency vs RPM plotted"


if __name__ == '__main__':
    from parapy.gui import display
    from parapy.geom import Position

    prop_system = HorizontalPropulsion(
        radius=1.0,
        envelope_radius=0.5,
        strut_size=0.1,
        position=Position(),
        horizontal_num_blades=3,
        c_tip=0.08,
        sweep_TE=5.0,
        airfoil_file='whitcomb.dat',
        blade_data_file_horizontal='blade_data_output2.xlsx',
        label='horizontalpropulsion'
    )
    display(prop_system)